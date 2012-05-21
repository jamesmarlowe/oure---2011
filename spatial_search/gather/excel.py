######################################################################
## @file excel.py
## @author James Marlowe - Class:Comp Sci 253 - Section:1A
## @brief This file writes data to excel files
######################################################################

from django.utils.encoding import smart_unicode
from openpyxl.cell import get_column_letter
from openpyxl import Workbook

## @fn writeToFile
## @brief writes a dictionary list to an excel file
## @pre user has a list of dictionaries to write to an excel file
## @post the list of dictionaries has been written to an excel file and saved
## @param dictionary: a list of dictionaries
## @param fname: the name of the excel file to be written
## @return nothing

def writeToFile(dictionary,fname):


    wb = Workbook()
    dest_filename = str(fname)+'.xlsx' #name the file
    print dest_filename
    ws = wb.worksheets[0]
    ws.title = "sorting_times"         #name the worksheet

    row = 1
    col_idx=1
    for key in dictionary[0].iterkeys():
        col = get_column_letter(col_idx)
        ws.cell('%s%s'%(col, row)).value = smart_unicode(key)    #save data to cell
        col_idx +=1

    row = 2
    for item in dictionary:            #loop through dictionary list
        col_idx=1
        for key in item.iterkeys():    #loop through individual dictionary items
            col = get_column_letter(col_idx)
            ws.cell('%s%s'%(col, row)).value = smart_unicode(item[key])    #save data to cell
            col_idx +=1
        row +=1

    wb.save(filename = dest_filename)
    return
    